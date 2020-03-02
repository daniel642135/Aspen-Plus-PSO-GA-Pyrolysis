import numpy as np
import deap
import random, operator, math
from deap import base
from deap import benchmarks
from deap import creator
from deap import tools
from own_package.others import create_excel_file, print_df_to_excel, print_array_to_excel
import openpyxl
import pickle
import pandas as pd
from deap.algorithms import varAnd
from deap.tools.selection import selRandom
import time

def generate_part(dim, pmin, pmax, smin, smax, int_idx):
    int_mask = [0]*dim
    # If int_idx is a list, int_mask will indicate which dimensions are integers
    try:
        for i in int_idx:
            int_mask[i] = 1
    except TypeError:
        pass  # If int_idx is None, then int_mask will be all 0
    position = [random.uniform(pmin[idx], pmax[idx])if int_mask[idx]==0 else random.randint(pmin[idx], pmax[idx]) for idx in range(dim)]
    part = creator.Particle(position)
    part.int_mask = int_mask
    part.speed = [random.uniform(smin[idx], smax[idx]) for idx in range(dim)]
    part.pmin = pmin[:]
    part.pmax = pmax[:]
    try:
        for single_int_idx in int_idx:
            # To make the lb and ub have equal probability of being selected after applying round
            part.pmin[single_int_idx] = pmin[single_int_idx] - 0.499
            part.pmax[single_int_idx] = pmax[single_int_idx] + 0.499
    except TypeError:
        pass
    part.smin = smin
    part.smax = smax
    return part

def updateParticle(part, best, w, c1, c2):
    r1 = (random.uniform(0, 1) for _ in range(len(part)))
    r2 = (random.uniform(0, 1) for _ in range(len(part)))
    v_u1 = map(operator.mul, [c1 for _ in range(len(part))], map(operator.mul, r1, map(operator.sub, part.best, part)))
    v_u2 = map(operator.mul, [c2 for _ in range(len(part))], map(operator.mul, r2, map(operator.sub, best, part)))
    v_inertia = map(operator.mul, [w for _ in range(len(part))], part.speed)
    part.speed = list(map(operator.add, v_inertia, map(operator.add, v_u1, v_u2)))
    for i, speed in enumerate(part.speed):
        if abs(speed) < part.smin[i]:
            part.speed[i] = math.copysign(part.smin[i], speed)
        elif abs(speed) > part.smax[i]:
            part.speed[i] = math.copysign(part.smax[i], speed)
    pos = list(map(operator.add, part, part.speed))
    for idx, (p, lb, ub, int_idx) in enumerate(zip(pos, part.pmin, part.pmax, part.int_mask)):
        pos[idx] = min(max(p, lb), ub)
        '''
        if p<lb:
            pos[idx] = lb
        if p > ub:
            pos[idx] = ub
        '''
        if int_idx == 1:
            pos[idx] = round(pos[idx])
    part[:] = pos

def ga_hybrid_polymutate(individual, eta, low, up, indpb):
    size = len(individual)

    for i, xl, xu, mask in zip(range(size), low, up, individual.int_mask):
        if random.random() <= indpb:
            x = individual[i]
            delta_1 = (x - xl) / (xu - xl)
            delta_2 = (xu - x) / (xu - xl)
            rand = random.random()
            mut_pow = 1.0 / (eta + 1.)

            if rand < 0.5:
                xy = 1.0 - delta_1
                val = 2.0 * rand + (1.0 - 2.0 * rand) * xy**(eta + 1)
                delta_q = val**mut_pow - 1.0
            else:
                xy = 1.0 - delta_2
                val = 2.0 * (1.0 - rand) + 2.0 * (rand - 0.5) * xy**(eta + 1)
                delta_q = 1.0 - val**mut_pow

            x = x + delta_q * (xu - xl)
            x = min(max(x, xl), xu)
            if mask == 1:
                x = round(x)
            individual[i] = x
    return individual,

def ga_hybrid_gaussianmutate(individual, sigma, low, up, indpb):
    size = len(individual)

    for i, xl, xu, s, mask in zip(range(size), low, up, sigma, individual.int_mask):
        if random.random() <= indpb:
            x = individual[i] + random.gauss(0, s)
            x = min(max(x, xl), xu)
            if mask == 1:
                x = round(x)
            individual[i] = x
    return individual,


def eval_func(part):
    return (sum(part),)

def pso_ga(func, pmin, pmax, smin, smax, int_idx, params, ga):
    # Setting params
    c1, c2, wmin, wmax, ga_iter_min, ga_iter_max, iter_gamma, ga_num_min, ga_num_max, num_beta,\
    tourn_size, cxpb, mutpb, indpd, eta,\
    pso_iter, swarm_size = \
    params['c1'], params['c2'], params['wmin'], params['wmax'],\
    params['ga_iter_min'], params['ga_iter_max'], params['iter_gamma'],\
    params['ga_num_min'], params['ga_num_max'], params['num_beta'],\
    params['tourn_size'], params['cxpd'], params['mutpd'], params['indpd'], params['eta'],\
    params['pso_iter'], params['swarm_size']

    # int_idx must be a list. If a single number is given, convert to list.
    if isinstance(int_idx,int):
        int_idx = [int_idx]

    creator.create("FitnessMin", base.Fitness, weights=(-1.0,))  # Minimization of a single scalar value
    creator.create("Particle", list, fitness=creator.FitnessMin, speed=list,
                   smin=None, smax=None, best=None, int_idx=None)

    toolbox = base.Toolbox()
    toolbox.register("particle", generate_part, dim=len(pmin), pmin=pmin, pmax=pmax, smin=smin, smax=smax, int_idx=int_idx)
    toolbox.register("population", tools.initRepeat, list, toolbox.particle)
    toolbox.register("update", updateParticle, c1=c1, c2=c2)
    toolbox.register("evaluate", func)

    toolbox.register("mate", tools.cxTwoPoint)
    #toolbox.register("mutate", ga_hybrid_polymutate, low=pmin, up=pmax, indpb=indpd, eta=eta)
    toolbox.register("mutate", ga_hybrid_gaussianmutate, low=pmin, up=pmax, indpb=indpd, sigma=smax)

    pop = toolbox.population(n=swarm_size)
    stats = tools.Statistics(lambda ind: ind.fitness.values)
    stats.register("avg", np.mean)
    stats.register("std", np.std)
    stats.register("min", np.min)
    stats.register("max", np.max)

    logbook = tools.Logbook()
    logbook.header = ["gen", "evals"] + stats.fields

    best = None
    pso_hof_num = max(1,round(ga_num_min*0.2))
    pso_hof = tools.HallOfFame(pso_hof_num)

    for g in range(pso_iter):
        # PSO segment first
        for part in pop:
            part.fitness.values = toolbox.evaluate(part)
            # Note: Fitness comparisons will compare the weighted value. Since weight is negative,
            # the comparison would be opposite unless you specify .values instead.
            if not part.best or part.best.fitness.values[0] > part.fitness.values[0]:
                part.best = creator.Particle(part)
                part.best.fitness.values = part.fitness.values
            if not best or best.fitness.values[0] > part.fitness.values[0]:
                best = creator.Particle(part)
                best.fitness.values = part.fitness.values
            #time.sleep(1)
        for part in pop:
            # Linear annealing for inertia velocity coefficient (the w weights)
            toolbox.update(part, best=best, w=wmax - (wmax-wmin)*g/pso_iter)
            #time.sleep(1)
        if ga:
            # GA segment
            # Start at max and approach min
            ga_pop = round(ga_num_min + (g/pso_iter)**num_beta*(ga_num_max-ga_iter_min))
            ga_gen = round(ga_iter_min + (g/pso_iter)**iter_gamma*(ga_iter_max-ga_iter_min))
            if len(pso_hof) == 0:
                ga_mask = [1 for _ in range(ga_pop)] + [0 for _ in range(swarm_size-ga_pop)]
                random.shuffle(ga_mask)
                population = [x for x,mask in zip(pop, ga_mask) if mask == 1]
            else:
                ga_pop += - pso_hof_num
                ga_mask = [1 for _ in range(ga_pop)] + [0 for _ in range(swarm_size - ga_pop)]
                random.shuffle(ga_mask)
                population = [x for x, mask in zip(pop, ga_mask) if mask == 1] + pso_hof.items

            halloffame = tools.HallOfFame(ga_pop)
            halloffame.update(population)
            ga_eval = 0
            # Begin the generational process
            for gen in range(ga_gen):
                # Select the next generation individuals. Built in tournament selector does not work for multi-objective
                # offspring = toolbox.select(population, len(population))
                # Own selection using tournment. Will work for multi-objective.
                chosen = []
                for i in range(ga_pop):
                    aspirants = selRandom(population, tourn_size)
                    scores = [x.fitness.values[0] for x in aspirants]
                    f = lambda i: scores[i]
                    chosen_idx = min(range(len(scores)), key=f)
                    chosen.append(aspirants[chosen_idx])
                    pass
                offspring = chosen

                # Vary the pool of individuals
                offspring = varAnd(offspring, toolbox, cxpb, mutpb)

                # Evaluate the individuals with an invalid fitness
                invalid_ind = [ind for ind in offspring if not ind.fitness.valid]
                ga_eval += len(invalid_ind)
                fitnesses = toolbox.map(toolbox.evaluate, invalid_ind)
                for ind, fit in zip(invalid_ind, fitnesses):
                    ind.fitness.values = fit

                # Update the hall of fame with the generated individuals
                halloffame.update(offspring)

                # Replace the current population by the offspring
                population[:] = offspring

            counter = 0
            if best.fitness.values[0] > halloffame[0].fitness.values[0]:
                best = creator.Particle(halloffame[0])
                best.fitness.values = halloffame[0].fitness.values
            for idx, mask in enumerate(ga_mask):
                if mask == 1:
                    try:
                        if pop[idx].fitness.values[0] > halloffame[counter].fitness.values[0]:
                            pop[idx] = halloffame[counter]
                            pop[idx].best = creator.Particle(part)
                            pop[idx].best.fitness.values = halloffame[counter].fitness.values
                        counter += 1
                    except IndexError:
                        break
        pso_hof.update(pop)

        # Gather all the fitnesses in one list and print the stats
        try:
            logbook.record(gen=g, evals=len(pop) + ga_eval, **stats.compile(pop))
        except UnboundLocalError:
            # Means ga=False and ga_eval is not assigned
            logbook.record(gen=g, evals=len(pop), **stats.compile(pop))
        #print(best)
        print(logbook.stream)

    print(best.fitness.values)

    # Printing to excel
    write_excel = create_excel_file('./results/pso_ga_results.xlsx')
    wb = openpyxl.load_workbook(write_excel)
    ws = wb[wb.sheetnames[-1]]

    ws.cell(1, 1).value = 'Optimal Decision Values'
    print_array_to_excel(['inlettemp', 'catalystweight', 'residencetime', 'reactorP'],(2,1), ws=ws, axis=1)
    print_array_to_excel(best, (3,1), ws=ws, axis=1)

    genfit = logbook.select("gen")
    avgfit = logbook.select("avg")
    stdfit = logbook.select("std")
    minfit = logbook.select("min")
    maxfit = logbook.select("max")

    ws.cell(5, 1).value = 'gen'
    ws.cell(6, 1).value = 'avg'
    ws.cell(7, 1).value = 'std'
    ws.cell(8, 1).value = 'min'
    ws.cell(9, 1).value = 'max'

    print_array_to_excel(genfit, (5,2), ws=ws, axis=1)
    print_array_to_excel(avgfit, (6, 2), ws=ws, axis=1)
    print_array_to_excel(stdfit, (7, 2), ws=ws, axis=1)
    print_array_to_excel(minfit, (8, 2), ws=ws, axis=1)
    print_array_to_excel(maxfit, (9, 2), ws=ws, axis=1)

    wb.save(write_excel)

    return pop, logbook, best


if __name__ == "__main__":
    params = {'c1': 1.5, 'c2':1.5, 'wmin': 0.4, 'wmax': 0.9,
              'ga_iter_min': 5, 'ga_iter_max': 20, 'iter_gamma': 10,
              'ga_num_min': 10, 'ga_num_max': 20, 'num_beta': 15,
              'tourn_size':3, 'cxpd': 0.5, 'mutpd':0.05, 'indpd': 0.5, 'eta':0.5,
              'pso_iter':200, 'swarm_size': 50}
    pmin = [-10,-10]
    pmax = [10,10]
    smin = [abs(x-y)*0.01 for x,y in zip(pmin, pmax)]
    smax = [abs(x-y)*0.5 for x,y in zip(pmin, pmax)]
    pso_ga(func=benchmarks.rosenbrock,pmin=pmin,pmax=pmax,
           smin=smin, smax=smax,
           int_idx=[1], params=params, ga=True)
